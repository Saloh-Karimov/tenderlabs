"""Golden tests for the pure in-memory conversion core.

The fixture CSV mirrors the Bluebeam Revu export column layout that
data_engine.py v2.1 codes against:
Subject, Length, Length Unit, Count, Page Label, Rise, Rise Unit, Code, Description
"""

import csv
import io
import zipfile

import openpyxl
import pytest

from app.core import processor
from app.core.processor import ParseError, convert

HEADER = [
    "Subject", "Length", "Length Unit", "Count",
    "Page Label", "Rise", "Rise Unit", "Code", "Description",
]

SAMPLE_ROWS = [
    # LEVEL 1: pipe lengths in metres, consolidated by code
    ["Pipe CHW", "12.5", "m", "1", "LEVEL 1", "", "", "P100", "CHW Pipe 100"],
    ["Pipe CHW", "7.5", "m", "1", "LEVEL 1", "", "", "P100", "CHW Pipe 100"],
    # PAIR subject doubles the count
    ["Valve PAIR", "", "", "3", "LEVEL 1", "", "", "V200", "Isolation Valve"],
    ["Elbow", "", "", "4", "LEVEL 1", "", "", "E300", "Elbow 90"],
    # No code/desc: counted in the raw total, never billed
    ["Note", "", "", "2", "LEVEL 1", "", "", "", ""],
    # Blank page label falls back to LEVEL 1
    ["Pipe CHW", "5", "m", "1", "", "", "", "P100", "CHW Pipe 100"],
    # LEVEL 2
    ["Pipe CHW", "10", "m", "1", "LEVEL 2", "", "", "P100", "CHW Pipe 100"],
    ["Elbow", "", "", "2.0", "LEVEL 2", "", "", "E300", "Elbow 90"],
]


def make_csv(rows) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(HEADER)
    writer.writerows(rows)
    # utf-8-sig BOM, exactly as Bluebeam exports
    return b"\xef\xbb\xbf" + buf.getvalue().encode("utf-8")


def zip_entries(result):
    zf = zipfile.ZipFile(io.BytesIO(result.zip_bytes))
    return zf, sorted(zf.namelist())


def load_sheet(zf, name, sheet):
    wb = openpyxl.load_workbook(io.BytesIO(zf.read(name)))
    return wb[sheet]


def rows_of(ws):
    return [[c.value for c in row] for row in ws.iter_rows()]


def test_level_by_level_golden():
    result = convert([("Tower A - CHW.csv", make_csv(SAMPLE_ROWS))],
                     "Tower A", lump_mode=False)

    assert result.system_count == 1
    assert result.row_count == 8
    assert result.warning_count == 0
    assert result.file_count == 3  # two levels + summary

    zf, names = zip_entries(result)
    assert names == [
        "CHW/Tower A - LEVEL 1 - CHW.xlsx",
        "CHW/Tower A - LEVEL 2 - CHW.xlsx",
        "Tower A - SUMMARY.xlsx",
    ]

    l1 = rows_of(load_sheet(zf, "CHW/Tower A - LEVEL 1 - CHW.xlsx", "Cavsoft_Import"))
    assert l1[0] == ["Code", "Description", "Units", "Quantity", "Target Section"]
    # 12.5 + 7.5 + 5 (blank label -> LEVEL 1) = 25, whole -> int
    assert l1[1] == ["P100", "CHW Pipe 100", "m", 25, "LEVEL 1"]
    # PAIR doubles 3 -> 6
    assert l1[2] == ["V200", "Isolation Valve", "ea", 6, "LEVEL 1"]
    assert l1[3] == ["E300", "Elbow 90", "ea", 4, "LEVEL 1"]
    assert len(l1) == 4

    v1 = rows_of(load_sheet(zf, "CHW/Tower A - LEVEL 1 - CHW.xlsx", "Verification"))
    # raw includes the code-less row (count 2): 12.5+7.5+6+4+2+5 = 37
    assert v1[0][1] == 37.0
    assert v1[1][1] == 35.0
    assert v1[2][1] == "*** NO ***"

    l2 = rows_of(load_sheet(zf, "CHW/Tower A - LEVEL 2 - CHW.xlsx", "Cavsoft_Import"))
    assert l2[1] == ["P100", "CHW Pipe 100", "m", 10, "LEVEL 2"]
    assert l2[2] == ["E300", "Elbow 90", "ea", 2, "LEVEL 2"]

    v2 = rows_of(load_sheet(zf, "CHW/Tower A - LEVEL 2 - CHW.xlsx", "Verification"))
    assert v2[0][1] == 12.0
    assert v2[1][1] == 12.0
    assert v2[2][1] == "YES"


def test_lump_sum_merges_everything():
    result = convert([("Tower A - CHW.csv", make_csv(SAMPLE_ROWS))],
                     "Tower A", lump_mode=True)

    zf, names = zip_entries(result)
    assert names == [
        "CHW/Tower A - LUMP SUM - CHW.xlsx",
        "Tower A - SUMMARY.xlsx",
    ]

    ws = rows_of(load_sheet(zf, "CHW/Tower A - LUMP SUM - CHW.xlsx", "Cavsoft_Import"))
    # P100 across all levels: 12.5+7.5+5+10 = 35
    assert ws[1] == ["P100", "CHW Pipe 100", "m", 35, "LUMP SUM"]
    assert ws[2] == ["V200", "Isolation Valve", "ea", 6, "LUMP SUM"]
    # E300: 4 + 2 = 6
    assert ws[3] == ["E300", "Elbow 90", "ea", 6, "LUMP SUM"]

    v = rows_of(load_sheet(zf, "CHW/Tower A - LUMP SUM - CHW.xlsx", "Verification"))
    assert v[0][1] == 49.0
    assert v[1][1] == 47.0


def test_master_summary_totals():
    result = convert([("chw export.csv", make_csv(SAMPLE_ROWS))],
                     "Tower A", lump_mode=False)
    zf, _ = zip_entries(result)
    ws = rows_of(load_sheet(zf, "Tower A - SUMMARY.xlsx", "Summary"))

    assert ws[0][0] == "Project: Tower A"
    assert ws[3] == ["System", "Level", "Raw", "Cavsoft", "Match?"]
    assert ws[4][0] == "CHW"
    assert ws[5][:5] == ["CHW", "LEVEL 1", 37.0, 35.0, "*** NO ***"]
    assert ws[6][:5] == ["CHW", "LEVEL 2", 12.0, 12.0, "YES"]
    assert ws[7][:5] == ["CHW TOTAL", None, 49.0, 47.0, "*** NO ***"]


def test_fractional_quantity_stays_float():
    rows = [["Pipe", "3.25", "m", "1", "LEVEL 1", "", "", "P1", "Pipe"]]
    result = convert([("hw.csv", make_csv(rows))], "T", lump_mode=False)
    zf, _ = zip_entries(result)
    ws = rows_of(load_sheet(zf, "HW/T - LEVEL 1 - HW.xlsx", "Cavsoft_Import"))
    assert ws[1][3] == 3.25


def test_system_detection_and_default():
    assert processor.detect_system("Job - SUPPLEMENTARY CW.csv") == "SUPPLEMENTARY CW"
    assert processor.detect_system("chilled chw riser.csv") == "CHW"
    assert processor.detect_system("mystery.csv") == ""

    rows = [["Pipe", "1", "m", "1", "LEVEL 1", "", "", "P1", "Pipe"]]
    result = convert([("mystery.csv", make_csv(rows))], "T", lump_mode=False)
    _, names = zip_entries(result)
    assert f"{processor.DEFAULT_SYSTEM}/T - LEVEL 1 - {processor.DEFAULT_SYSTEM}.xlsx" in names


def test_tender_name_is_sanitised_in_filenames():
    rows = [["Pipe", "1", "m", "1", "LEVEL 1", "", "", "P1", "Pipe"]]
    result = convert([("chw.csv", make_csv(rows))], 'To/wer: "A"', lump_mode=False)
    _, names = zip_entries(result)
    assert names == [
        "CHW/To-wer- -A- - LEVEL 1 - CHW.xlsx",
        "To-wer- -A- - SUMMARY.xlsx",
    ]


def test_levels_without_identifiers_are_skipped_and_counted():
    rows = [
        ["Pipe", "1", "m", "1", "LEVEL 1", "", "", "P1", "Pipe"],
        ["Note", "", "", "5", "LEVEL 9", "", "", "", ""],
    ]
    result = convert([("chw.csv", make_csv(rows))], "T", lump_mode=False)
    assert result.warning_count == 1
    _, names = zip_entries(result)
    assert "CHW/T - LEVEL 1 - CHW.xlsx" in names
    assert not any("LEVEL 9" in n for n in names)


def test_empty_and_invalid_inputs_raise_parse_errors():
    with pytest.raises(ParseError):
        convert([("a.csv", make_csv([]))], "T", lump_mode=False)  # header only

    with pytest.raises(ParseError):
        convert([("a.csv", b"")], "T", lump_mode=False)  # nothing at all

    with pytest.raises(ParseError):
        convert([], "T", lump_mode=False)  # no files at all

    with pytest.raises(ParseError):  # rows exist, none billable
        convert([("a.csv", make_csv([["Note", "", "", "5", "L1", "", "", "", ""]]))],
                "T", lump_mode=False)

    with pytest.raises(ParseError):  # tender name sanitises to nothing
        convert([("a.csv", make_csv([["Pipe", "1", "m", "1", "L1", "", "", "P1", "Pipe"]]))],
                "   ", lump_mode=False)


def test_batch_multi_system_single_zip_and_summary():
    chw = [["Pipe", "10", "m", "1", "LEVEL 1", "", "", "C1", "CHW Pipe"]]
    hw = [["Pipe", "4", "m", "1", "LEVEL 1", "", "", "H1", "HW Pipe"],
          ["Pipe", "6", "m", "1", "LEVEL 2", "", "", "H1", "HW Pipe"]]
    cw = [["Valve PAIR", "", "", "2", "ROOF", "", "", "W1", "CW Valve"]]

    result = convert(
        [("Job - CHW.csv", make_csv(chw)),
         ("Job - HW.csv", make_csv(hw)),
         ("Job - CW.csv", make_csv(cw))],
        "Job", lump_mode=False,
    )

    assert result.system_count == 3
    assert result.row_count == 4
    assert result.file_count == 5  # 1 CHW + 2 HW + 1 CW workbooks + summary

    zf, names = zip_entries(result)
    assert names == [
        "CHW/Job - LEVEL 1 - CHW.xlsx",
        "CW/Job - ROOF - CW.xlsx",
        "HW/Job - LEVEL 1 - HW.xlsx",
        "HW/Job - LEVEL 2 - HW.xlsx",
        "Job - SUMMARY.xlsx",
    ]

    ws = rows_of(load_sheet(zf, "Job - SUMMARY.xlsx", "Summary"))
    # One unified summary: CHW section, then HW, then CW, in upload order
    assert ws[4][0] == "CHW"
    assert ws[5][:5] == ["CHW", "LEVEL 1", 10.0, 10.0, "YES"]
    assert ws[6][:5] == ["CHW TOTAL", None, 10.0, 10.0, "YES"]
    assert ws[7][0] == "HW"
    assert ws[8][:5] == ["HW", "LEVEL 1", 4.0, 4.0, "YES"]
    assert ws[9][:5] == ["HW", "LEVEL 2", 6.0, 6.0, "YES"]
    assert ws[10][:5] == ["HW TOTAL", None, 10.0, 10.0, "YES"]
    assert ws[11][0] == "CW"
    assert ws[12][:5] == ["CW", "ROOF", 4.0, 4.0, "YES"]  # PAIR doubles 2 -> 4
    assert ws[13][:5] == ["CW TOTAL", None, 4.0, 4.0, "YES"]


def test_batch_same_system_files_merge_quantities():
    a = [["Pipe", "10", "m", "1", "LEVEL 1", "", "", "C1", "CHW Pipe"]]
    b = [["Pipe", "5", "m", "1", "LEVEL 1", "", "", "C1", "CHW Pipe"]]

    result = convert(
        [("Riser CHW.csv", make_csv(a)), ("Plant CHW.csv", make_csv(b))],
        "Job", lump_mode=False,
    )

    assert result.system_count == 1
    zf, names = zip_entries(result)
    # One workbook per level, no duplicate-name collisions
    assert names == ["CHW/Job - LEVEL 1 - CHW.xlsx", "Job - SUMMARY.xlsx"]

    ws = rows_of(load_sheet(zf, "CHW/Job - LEVEL 1 - CHW.xlsx", "Cavsoft_Import"))
    assert ws[1] == ["C1", "CHW Pipe", "m", 15, "LEVEL 1"]


def test_batch_error_names_offending_file():
    good = [["Pipe", "1", "m", "1", "L1", "", "", "P1", "Pipe"]]
    with pytest.raises(ParseError, match="empty HW.csv"):
        convert([("chw.csv", make_csv(good)), ("empty HW.csv", make_csv([]))],
                "T", lump_mode=False)


def test_core_module_never_touches_disk():
    """ZDR guard: the banned filesystem tokens must not appear in core."""
    import inspect

    src = inspect.getsource(processor)
    assert "tempfile" not in src
    assert "open(" not in src
