"""
TenderLabs conversion core.

Pure, in-memory port of the Ghost Architect data engine (v2.1):
Bluebeam Revu CSV bytes in -> CAVsoft-import XLSX ZIP bytes out.

ZERO DATA RETENTION: nothing in this module may touch the filesystem,
raise messages containing cell values, or log anything at all.
"""

import csv
import io
import re
import zipfile
from collections import OrderedDict
from dataclasses import dataclass

import openpyxl
from openpyxl.styles import Font, PatternFill

SYSTEMS = ["SUPPLEMENTARY CW", "TENANT CW", "AMBIENT LOOP", "CHW", "HW", "CW", "REF"]
DEFAULT_SYSTEM = "GENERAL"

# 1-based Bluebeam export column positions (per data_engine.py v2.1).
COL_SUBJECT    = 1
COL_LENGTH     = 2
COL_LEN_UNIT   = 3
COL_COUNT      = 4
COL_PAGE_LABEL = 5
COL_RISE       = 6
COL_RISE_UNIT  = 7
COL_CODE       = 8
COL_DESC       = 9

FILL_HEADER = PatternFill("solid", fgColor="DCDCDC")
FILL_SYSTEM = PatternFill("solid", fgColor="DCE6F1")
FILL_GRAND  = PatternFill("solid", fgColor="FFF2CC")
FONT_BOLD   = Font(bold=True)
FONT_GREEN  = Font(bold=True, color="009600")
FONT_RED    = Font(bold=True, color="C80000")
FONT_LINK_BOLD = Font(bold=True, color="0000FF", underline="single")
FONT_LINK      = Font(color="0000FF", underline="single")


class ParseError(Exception):
    """Input CSV cannot be processed. Messages may reference column headers
    or structure ONLY — never cell values."""


# Deliberately carries only zero-retention-safe metadata besides the ZIP
# payload itself: counts only, never row data, level names, or system tags.
@dataclass
class ConversionResult:
    zip_bytes: bytes
    system_count: int
    row_count: int
    file_count: int
    warning_count: int


def sanitise(name: str) -> str:
    for ch in r'/\:*?"<>|':
        name = name.replace(ch, "-")
    return name.strip()


def detect_system(filename: str) -> str:
    for s in SYSTEMS:
        if s.upper() in filename.upper():
            return s
    return ""


def to_float(v) -> float:
    try:
        f = float(v)
        return f if f == f else 0.0
    except (TypeError, ValueError):
        return 0.0


def col(row, c):
    idx = c - 1
    return row[idx] if idx < len(row) else None


def str_col(row, c) -> str:
    v = col(row, c)
    return str(v).strip() if v is not None else ""


def read_export(csv_bytes: bytes):
    """Decode a Bluebeam CSV export from bytes. Returns (header, rows)."""
    try:
        text = csv_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = csv_bytes.decode("latin-1")
        except UnicodeDecodeError as exc:
            raise ParseError("CSV is not valid UTF-8 or Latin-1 text.") from exc

    try:
        data = list(csv.reader(io.StringIO(text)))
    except csv.Error as exc:
        raise ParseError("CSV structure could not be parsed.") from exc

    header = data[0] if data else []
    rows = data[1:] if len(data) > 1 else []
    return header, rows


def consolidate_rows(level_rows, lump_sum_mode=False):
    consolidated = OrderedDict()
    level_name = "LUMP SUM" if lump_sum_mode else ""
    raw_total = 0.0
    cavsoft_total = 0.0

    for row in level_rows:
        code    = str_col(row, COL_CODE)
        desc    = str_col(row, COL_DESC)
        subject = str_col(row, COL_SUBJECT)
        level   = "LUMP SUM" if lump_sum_mode else (str_col(row, COL_PAGE_LABEL) or "LEVEL 1")
        length  = to_float(col(row, COL_LENGTH))
        count   = to_float(col(row, COL_COUNT))

        if not lump_sum_mode:
            level_name = level

        is_pair = "PAIR" in subject.upper()

        if length > 0:
            qty = length
            units = "m"
        else:
            qty = (count * 2) if is_pair else count
            units = "ea"

        raw_total += qty

        # QA protocol: reject rows missing database identifiers.
        if not code or not desc:
            continue

        if code not in consolidated:
            consolidated[code] = {"desc": desc, "units": units, "qty": 0.0, "level": level}

        consolidated[code]["qty"] += qty
        cavsoft_total += qty

    cavsoft_rows = []
    for code, d in consolidated.items():
        qty = round(d["qty"], 3)
        qty_out = int(qty) if qty == int(qty) else qty
        cavsoft_rows.append([code, d["desc"], d["units"], qty_out, d["level"]])

    return cavsoft_rows, round(raw_total, 3), round(cavsoft_total, 3), level_name


def build_level_workbooks(rows, system_name, tender_name, lump_sum_mode=False):
    """Group rows by level (or a single LUMP SUM node) and build one
    CAVsoft-import workbook per level, entirely in memory.

    Returns (results, skipped_count) where each result carries the level
    name, the in-zip filename, xlsx bytes, and verification totals.
    """
    level_map = OrderedDict()

    if lump_sum_mode:
        level_map["LUMP SUM"] = rows
    else:
        for row in rows:
            level = str_col(row, COL_PAGE_LABEL) or "LEVEL 1"
            if level not in level_map:
                level_map[level] = []
            level_map[level].append(row)

    results = []
    skipped = 0

    for level, l_rows in level_map.items():
        cavsoft_rows, raw_total, cav_total, _ = consolidate_rows(l_rows, lump_sum_mode)

        # Page scrubber: drop levels with no code/desc-bearing rows.
        if not cavsoft_rows:
            skipped += 1
            continue

        clean_lvl = sanitise(level)
        filename = f"{tender_name} - {clean_lvl} - {system_name}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cavsoft_Import"

        col_headers = ["Code", "Description", "Units", "Quantity", "Target Section"]
        for c_idx, h in enumerate(col_headers, 1):
            c = ws.cell(row=1, column=c_idx, value=h)
            c.font = FONT_BOLD
            c.fill = FILL_HEADER

        for r_idx, cav_row in enumerate(cavsoft_rows, start=2):
            for c_idx, val in enumerate(cav_row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        vws = wb.create_sheet("Verification")
        vws["A1"] = "Raw Bluebeam Total (All)"
        vws["B1"] = raw_total
        vws["A2"] = "Cavsoft Billed Total"
        vws["B2"] = cav_total

        match = round(raw_total, 2) == round(cav_total, 2)
        vws["A3"] = "Match?"
        vws["B3"] = "YES" if match else "*** NO ***"
        vws["B3"].font = FONT_GREEN if match else FONT_RED

        for r in range(1, 4):
            vws.cell(row=r, column=1).font = FONT_BOLD

        vws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["E"].width = 20

        buf = io.BytesIO()
        wb.save(buf)
        wb.close()

        results.append({
            "level": level,
            "filename": filename,
            "xlsx_bytes": buf.getvalue(),
            "raw_total": raw_total,
            "cav_total": cav_total,
        })

    return results, skipped


def build_master_summary(tender_name, system_results) -> bytes:
    """Build the master Summary workbook in memory. Hyperlinks are relative
    paths into the extracted ZIP so they resolve once unzipped."""
    mwb = openpyxl.Workbook()
    mws = mwb.active
    mws.title = "Summary"

    mws["A1"] = f"Project: {tender_name}"
    mws["A1"].font = FONT_BOLD
    mws["A2"] = "GRAND TOTAL"
    mws["C2"] = '=SUMIF(A5:A100000,"* TOTAL",C5:C100000)'
    mws["D2"] = '=SUMIF(A5:A100000,"* TOTAL",D5:D100000)'
    mws["E2"] = '=IF(ROUND(C2,2)=ROUND(D2,2),"YES","*** NO ***")'

    for c in mws["A2:E2"][0]:
        c.font = FONT_BOLD
        c.fill = FILL_GRAND

    headers = ["System", "Level", "Raw", "Cavsoft", "Match?"]
    for c_idx, val in enumerate(headers, 1):
        c = mws.cell(row=4, column=c_idx, value=val)
        c.font = FONT_BOLD
        c.fill = FILL_HEADER

    cur_r = 5
    for system_name, levels in system_results:
        sc = mws.cell(row=cur_r, column=1, value=system_name)
        sc.fill = FILL_SYSTEM
        sc.font = FONT_LINK_BOLD
        sc.hyperlink = f"{system_name}/"
        cur_r += 1

        sys_raw = 0.0
        sys_cav = 0.0
        for lvl in levels:
            mws.cell(row=cur_r, column=1, value=system_name)
            lc = mws.cell(row=cur_r, column=2, value=lvl["level"])
            lc.font = FONT_LINK
            lc.hyperlink = f"{system_name}/{lvl['filename']}"

            mws.cell(row=cur_r, column=3, value=round(lvl["raw_total"], 2))
            mws.cell(row=cur_r, column=4, value=round(lvl["cav_total"], 2))

            match = round(lvl["raw_total"], 2) == round(lvl["cav_total"], 2)
            mws_match = mws.cell(row=cur_r, column=5, value="YES" if match else "*** NO ***")
            mws_match.font = FONT_GREEN if match else FONT_RED

            sys_raw += lvl["raw_total"]
            sys_cav += lvl["cav_total"]
            cur_r += 1

        tc = mws.cell(row=cur_r, column=1, value=f"{system_name} TOTAL")
        tc.font = FONT_BOLD
        mws.cell(row=cur_r, column=3, value=round(sys_raw, 2)).font = FONT_BOLD
        mws.cell(row=cur_r, column=4, value=round(sys_cav, 2)).font = FONT_BOLD

        sys_match = round(sys_raw, 2) == round(sys_cav, 2)
        tc_match = mws.cell(row=cur_r, column=5, value="YES" if sys_match else "*** NO ***")
        tc_match.font = FONT_GREEN if sys_match else FONT_RED
        cur_r += 1

    mws.column_dimensions["A"].width = 22
    mws.column_dimensions["B"].width = 45
    mws.column_dimensions["C"].width = 14

    buf = io.BytesIO()
    mwb.save(buf)
    mwb.close()
    return buf.getvalue()


def convert(files: list[tuple[str, bytes]], tender_name: str,
            lump_mode: bool) -> ConversionResult:
    """Batch pipeline: a list of (source filename, CSV bytes) exports ->
    one ZIP of CAVsoft imports plus a single unified master summary.

    Mirrors the prototype's export_paths loop: the system is detected per
    file from its filename. Files that map to the same system are merged
    before level grouping, so their quantities consolidate together.
    """
    tender_name = sanitise(tender_name)
    if not tender_name:
        raise ParseError("Tender name is empty after sanitising.")
    if not files:
        raise ParseError("No files provided.")

    system_rows: OrderedDict[str, list] = OrderedDict()
    row_count = 0

    for source_filename, csv_bytes in files:
        system_name = detect_system(source_filename) or DEFAULT_SYSTEM
        try:
            _header, rows = read_export(csv_bytes)
        except ParseError as exc:
            raise ParseError(f"'{source_filename}': {exc}") from exc
        if not rows:
            raise ParseError(
                f"'{source_filename}' contains a header but no data rows."
            )
        system_rows.setdefault(system_name, []).extend(rows)
        row_count += len(rows)

    system_results = []
    workbook_count = 0
    skipped_total = 0

    for system_name, rows in system_rows.items():
        results, skipped = build_level_workbooks(
            rows, system_name, tender_name, lump_sum_mode=lump_mode
        )
        skipped_total += skipped
        if results:
            system_results.append((system_name, results))
            workbook_count += len(results)

    if not system_results:
        raise ParseError(
            "No rows have both the Code and Description columns populated "
            f"(expected in columns {COL_CODE} and {COL_DESC})."
        )

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for system_name, results in system_results:
            for r in results:
                zf.writestr(f"{system_name}/{r['filename']}", r["xlsx_bytes"])
        summary_bytes = build_master_summary(tender_name, system_results)
        zf.writestr(f"{tender_name} - SUMMARY.xlsx", summary_bytes)

    return ConversionResult(
        zip_bytes=zip_buf.getvalue(),
        system_count=len(system_results),
        row_count=row_count,
        file_count=workbook_count + 1,
        warning_count=skipped_total,
    )
