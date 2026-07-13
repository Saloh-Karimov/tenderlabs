"""
===============================================================
GHOST ARCHITECT: DATA ENGINE (V2.1 - CLEAN ARCHITECTURE)
===============================================================
Backend module handling dual-mode Data Transformation. 
Generates strict CAVsoft imports (Level-by-Level OR Lump Sum).
Fully Unix/Mac/Windows compliant.
"""

import os
import sys
import re
from collections import OrderedDict
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox

# ===============================================================
# GLOBAL CONFIG & STYLES
# ===============================================================
BASE_MARKUPS = os.path.expanduser("~/Desktop/Markups") if sys.platform == "darwin" else r"C:\Users\Owner\OneDrive - MCM Mechanical Services\Desktop\Markups"
SYSTEMS = ["SUPPLEMENTARY CW", "TENANT CW", "AMBIENT LOOP", "CHW", "HW", "CW", "REF"]

COL_SUBJECT    = 1
COL_LENGTH     = 2
COL_LEN_UNIT   = 3
COL_COUNT      = 4
COL_PAGE_LABEL = 5
COL_RISE       = 6
COL_RISE_UNIT  = 7
COL_CODE       = 8
COL_DESC       = 9

FILL_HEADER = PatternFill("solid", fgColor="DCDCDC")
FILL_SYSTEM = PatternFill("solid", fgColor="DCE6F1")
FILL_GRAND  = PatternFill("solid", fgColor="FFF2CC")
FONT_BOLD   = Font(bold=True)
FONT_GREEN  = Font(bold=True, color="009600")
FONT_RED    = Font(bold=True, color="C80000")
FONT_LINK_BOLD = Font(bold=True, color="0000FF", underline="single")
FONT_LINK      = Font(color="0000FF", underline="single")


# ===============================================================
# UTILITIES & TRANSLATION MATRICES
# ===============================================================
def sanitise(name: str) -> str:
    for ch in r'/\:*?"<>|': 
        name = name.replace(ch, "-")
    return name.strip()

def detect_system(filename: str) -> str:
    for s in SYSTEMS:
        if s.upper() in filename.upper(): 
            return s
    return ""

def to_float(v) -> float:
    try:
        f = float(v)
        return f if f == f else 0.0
    except (TypeError, ValueError): 
        return 0.0

def read_export(path: str):
    ext = Path(path).suffix.lower()
    if ext == ".csv":
        import csv
        with open(path, newline="", encoding="utf-8-sig") as f: 
            data = list(csv.reader(f))
    else:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        data = [[cell.value for cell in row] for row in ws.iter_rows()]
        wb.close()
    
    header = data[0] if data else []
    rows = data[1:] if len(data) > 1 else []
    return header, rows

def col(row, c):
    idx = c - 1
    return row[idx] if idx < len(row) else None

def str_col(row, c) -> str:
    v = col(row, c)
    return str(v).strip() if v is not None else ""

def sort_levels(levels):
    def get_sort_key(l):
        lu = l.upper().strip()
        prio = 0
        if "LUMP SUM" in lu: 
            prio = 1000
        elif "ROOF" in lu: 
            prio = 100
        elif "LEVEL" in lu or "FIRST" in lu or "1ST" in lu: 
            prio = 80
        elif "GROUND" in lu or " G " in lu or lu == "G": 
            prio = 50
        elif "BASEMENT" in lu or lu.startswith("B-"): 
            prio = -50
        
        floor_num = 0
        f_match = re.search(r'LEVEL\s*(\d+)', lu)
        if f_match: 
            floor_num = int(f_match.group(1))
        elif "FIRST" in lu: 
            floor_num = 1
        
        build_num = 0
        b_match = re.search(r'B(\d+)', lu)
        if b_match: 
            build_num = int(b_match.group(1))
        
        return (prio, floor_num, build_num, lu)
        
    return sorted(list(levels), key=get_sort_key, reverse=True)


# ===============================================================
# DATA AGGREGATION & ARCHITECTURE LOGIC 
# ===============================================================
def consolidate_rows(level_rows, lump_sum_mode=False):
    consolidated = OrderedDict()
    level_name = "LUMP SUM" if lump_sum_mode else ""
    raw_total = 0.0
    cavsoft_total = 0.0

    for row in level_rows:
        code    = str_col(row, COL_CODE)
        desc    = str_col(row, COL_DESC)
        subject = str_col(row, COL_SUBJECT)
        # Override the level if running in bulk extraction mode
        level   = "LUMP SUM" if lump_sum_mode else (str_col(row, COL_PAGE_LABEL) or "LEVEL 1")
        length  = to_float(col(row, COL_LENGTH))
        count   = to_float(col(row, COL_COUNT))

        if not lump_sum_mode:
            level_name = level

        is_pair = "PAIR" in subject.upper()

        if length > 0:
            qty = length
            units = "m"
        else:
            qty = (count * 2) if is_pair else count
            units = "ea"

        raw_total += qty
        
        # Ghost QA Protocol: Reject blank database inputs
        if not code or not desc: 
            continue

        if code not in consolidated:
            consolidated[code] = {"desc": desc, "units": units, "qty": 0.0, "level": level}
            
        consolidated[code]["qty"] += qty
        cavsoft_total += qty

    cavsoft_rows = []
    for code, d in consolidated.items():
        qty = round(d["qty"], 3)
        qty_out = int(qty) if qty == int(qty) else qty
        cavsoft_rows.append([code, d["desc"], d["units"], qty_out, d["level"]])
        
    return cavsoft_rows, round(raw_total, 3), round(cavsoft_total, 3), level_name

def build_level_files(export_path, system_name, project_folder, tender_name, lump_sum_mode=False):
    headers, rows = read_export(export_path)
    level_map = OrderedDict()

    # Dynamic Dictionary Sorting Check
    if lump_sum_mode:
        virtual_level = "LUMP SUM"
        level_map[virtual_level] = rows  # Pipe everything to a single key natively
    else:
        for row in rows:
            level = str_col(row, COL_PAGE_LABEL) or "LEVEL 1"
            if level not in level_map: 
                level_map[level] = []
            level_map[level].append(row)

    if not level_map: 
        return []

    system_folder = os.path.join(project_folder, system_name)
    os.makedirs(system_folder, exist_ok=True)
    results = []

    for level, l_rows in level_map.items():
        cavsoft_rows, raw_total, cav_total, _ = consolidate_rows(l_rows, lump_sum_mode)
        
        # GHOST PAGE SCRUBBER 
        if not cavsoft_rows:
            print(f"    [!] Skipping node '{level}': Missing code/desc identifiers.")
            continue

        clean_lvl = sanitise(level)
        save_name = f"{tender_name} - {clean_lvl} - {system_name}.xlsx"
        save_path = os.path.join(system_folder, save_name)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cavsoft_Import"

        # Structural Injection
        col_headers = ["Code", "Description", "Units", "Quantity", "Target Section"]
        for c_idx, h in enumerate(col_headers, 1):
            c = ws.cell(row=1, column=c_idx, value=h)
            c.font = FONT_BOLD
            c.fill = FILL_HEADER
        
        for r_idx, cav_row in enumerate(cavsoft_rows, start=2):
            for c_idx, val in enumerate(cav_row, start=1): 
                ws.cell(row=r_idx, column=c_idx, value=val)

        # Integrity Checks
        vws = wb.create_sheet("Verification")
        vws["A1"] = "Raw Bluebeam Total (All)"
        vws["B1"] = raw_total
        vws["A2"] = "Cavsoft Billed Total"
        vws["B2"] = cav_total
        
        match = round(raw_total, 2) == round(cav_total, 2)
        vws["A3"] = "Match?"
        vws["B3"] = "YES" if match else "*** NO ***"
        vws["B3"].font = FONT_GREEN if match else FONT_RED
        
        for r in range(1, 4): 
            vws.cell(row=r, column=1).font = FONT_BOLD
            
        vws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["E"].width = 20

        wb.save(save_path)
        wb.close()
        results.append({"level": level, "save_path": save_path, "raw_total": raw_total, "cav_total": cav_total})

    return results

def write_master_summary(master_path, tender_name, system_results):
    if os.path.exists(master_path):
        mwb = openpyxl.load_workbook(master_path)
    else:
        mwb = openpyxl.Workbook()
        
    mws = mwb["Summary"] if "Summary" in mwb.sheetnames else mwb.active
    mws.title = "Summary"

    for row in mws.iter_rows(min_row=5):
        for cell in row: 
            cell.value = None
            cell.hyperlink = None

    if not mws["A1"].value:
        mws["A1"] = f"Project: {tender_name}"
        mws["A1"].font = FONT_BOLD
        mws["A2"] = "GRAND TOTAL"
        mws["C2"] = '=SUMIF(A5:A100000,"* TOTAL",C5:C100000)'
        mws["D2"] = '=SUMIF(A5:A100000,"* TOTAL",D5:D100000)'
        mws["E2"] = '=IF(ROUND(C2,2)=ROUND(D2,2),"YES","*** NO ***")'
        
        for c in mws["A2:E2"][0]: 
            c.font = FONT_BOLD
            c.fill = FILL_GRAND
            
        headers = ["System", "Level", "Raw", "Cavsoft", "Match?"]
        for c_idx, val in enumerate(headers, 1): 
            c = mws.cell(row=4, column=c_idx, value=val)
            c.font = FONT_BOLD
            c.fill = FILL_HEADER

    cur_r = 5
    for system_name, system_folder, levels in system_results:
        sc = mws.cell(row=cur_r, column=1, value=system_name)
        sc.fill = FILL_SYSTEM
        sc.font = FONT_LINK_BOLD
        
        try: 
            sc.hyperlink = system_folder
        except: 
            pass
            
        cur_r += 1

        sys_raw = 0.0
        sys_cav = 0.0
        for lvl in levels:
            mws.cell(row=cur_r, column=1, value=system_name)
            lc = mws.cell(row=cur_r, column=2, value=lvl["level"])
            lc.font = FONT_LINK
            
            try: 
                lc.hyperlink = lvl["save_path"]
            except: 
                pass
            
            mws.cell(row=cur_r, column=3, value=round(lvl["raw_total"], 2))
            mws.cell(row=cur_r, column=4, value=round(lvl["cav_total"], 2))

            match = round(lvl["raw_total"], 2) == round(lvl["cav_total"], 2)
            mws_match = mws.cell(row=cur_r, column=5, value="YES" if match else "*** NO ***")
            mws_match.font = FONT_GREEN if match else FONT_RED
            
            sys_raw += lvl["raw_total"]
            sys_cav += lvl["cav_total"]
            cur_r += 1

        tc = mws.cell(row=cur_r, column=1, value=f"{system_name} TOTAL")
        tc.font = FONT_BOLD
        mws.cell(row=cur_r, column=3, value=round(sys_raw, 2)).font = FONT_BOLD
        mws.cell(row=cur_r, column=4, value=round(sys_cav, 2)).font = FONT_BOLD
        
        sys_match = round(sys_raw, 2) == round(sys_cav, 2)
        tc_match = mws.cell(row=cur_r, column=5, value="YES" if sys_match else "*** NO ***")
        tc_match.font = FONT_GREEN if sys_match else FONT_RED
        cur_r += 1

    mws.column_dimensions["A"].width = 22
    mws.column_dimensions["B"].width = 45
    mws.column_dimensions["C"].width = 14
    
    mwb.save(master_path)
    mwb.close()

def print_terminal_summary(tender_name, system_results):
    print("\n" + "═" * 78)
    print(f"  DATA ENGINE METRICS ─ [{tender_name}]")
    print("═" * 78)

    g_raw = 0.0
    g_cav = 0.0
    for sys_name, _, lvls in system_results:
        print(f"\n  ▶ {sys_name}")
        print(f"  {'Level/Target Section':<35} {'Raw':>10}  {'Cavsoft':>10}  Match")
        s_raw = 0.0
        s_cav = 0.0
        
        for l in lvls:
            r = l['raw_total']
            c = l['cav_total']
            m = "✓" if round(r, 2) == round(c, 2) else "✗"
            print(f"  {l['level']:<35} {r:>10.2f}  {c:>10.2f}   {m}")
            s_raw += r
            s_cav += c
            g_raw += r
            g_cav += c
            
        print(f"  {'  SYSTEM TOTAL':<35} {s_raw:>10.2f}  {s_cav:>10.2f}")
    
    print("\n" + "═" * 78)
    print(f"  {'GLOBAL PAYLOAD TOTAL':<35} {g_raw:>10.2f}  {g_cav:>10.2f}")
    print("═" * 78)


# ===============================================================
# MAIN LOCAL GUI TESTING ENVIRONMENT
# ===============================================================
def main():
    # If running on macOS without standard display bridges, handle Tkinter elegantly
    try:
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        
        tender_name = simpledialog.askstring("Project Title", "Enter Project Name:", parent=root)
        if not tender_name: 
            return
            
        lump_response = messagebox.askyesno("Operational Override", "Merge data into single 'LUMP SUM' mode?", parent=root)
        lump_mode = True if lump_response else False
        
        print("\n[*] Upload payload (raw files)...")
        export_paths = filedialog.askopenfilenames(
            title="Select Raw Exports", 
            filetypes=[("Data Files", "*.csv *.xlsx *.xlsm *.xls"), ("All Files", "*.*")]
        )
        if not export_paths: 
            return
            
    except Exception as gui_e:
        print("\n[!] GUI Blocked (Running via non-display terminal). Engaging fallback input loop...")
        tender_name = input("Enter Project Title: ").strip()
        if not tender_name: return
        
        lm_resp = input("Merge data into single 'LUMP SUM' mode? (y/n): ").strip().lower()
        lump_mode = True if lm_resp == 'y' else False
        
        path_str = input("Drag and drop your raw exported CSV/XLSX file here (and press Enter): ").strip().strip("'\"")
        export_paths = [path_str] if os.path.exists(path_str) else []
        if not export_paths:
            print("File not found. Aborting.")
            return

    print("\n" + "="*50)
    print("  GHOST ARCHITECT - B2B BACKEND V2")
    print("="*50)

    if lump_mode: 
        print("\n[i] LUMP SUM OVERRIDE ENGAGED. Ignoring local level markers.")

    project_folder = os.path.join(BASE_MARKUPS, tender_name.strip())
    os.makedirs(project_folder, exist_ok=True)
    master_path = os.path.join(project_folder, f"{tender_name.strip()} - SUMMARY.xlsx")

    system_results = []
    
    for path in export_paths:
        # Handling edge case if Path inputs are wrapped weirdly by OS
        clean_path = str(Path(path))
        sys_name = detect_system(Path(clean_path).stem)
        
        if not sys_name:
            sys_name = input(f"\nTarget unreadable for {Path(clean_path).name}.\nEnter designation manually (e.g. CHW): ").strip()
            if not sys_name: 
                continue

        try:
            l_results = build_level_files(clean_path, sys_name, project_folder, tender_name, lump_sum_mode=lump_mode)
            if l_results: 
                system_results.append((sys_name, os.path.join(project_folder, sys_name), l_results))
        except Exception as e:
            print(f"[!] Critical block collapse {clean_path} | EXCEPTION: {e}")
            continue

    if not system_results: 
        print("\n[!] Zero mapped database records survived the purge check. Operation terminated.")
        sys.exit(0)
        
    write_master_summary(master_path, tender_name, system_results)
    print_terminal_summary(tender_name, system_results)
    
    print(f"\n[✔] Data transformation achieved locally. Offline backend operations cleared. View summary at:\n {master_path}\n")

if __name__ == "__main__":
    main()